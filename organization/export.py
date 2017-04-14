'''
Created on Nov 15, 2016

@author: nanda
'''
from django.contrib.auth.decorators import login_required
from django.http.response import HttpResponse
from organization.models import *
import csv
import xlwt
from xlwt.compat import xrange
from openpyxl.reader.excel import load_workbook
from organization.excelRow import insert_rows
from openpyxl.styles import Color, Fill
from cLIMS.base import WORKSPACEPATH, FILEUPLOADPATH
from dryLab.models import *
import json
from _collections import OrderedDict, defaultdict
from wetLab.models import *
import collections
from openpyxl.utils.cell import get_column_letter


@login_required 
def exportExperiment(request):
    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="meta_data_sample.csv"'
    projectId = request.session['projectId']
    dbdata = Experiment.objects.filter(project=projectId)
    writer = csv.writer(response)
    for a in dbdata:
        field_names = [x.name for x in a._meta.local_fields]
        break
    writer.writerow(field_names)
    for obj in dbdata:
        writer.writerow([getattr(obj, field) for field in field_names])
    return response

def orderByNumber(jsonDict):
    jsonList = jsonDict.items()
    sorted_list = sorted(jsonList, key=lambda k: (int(k[1]['order'])))
    sorted_dict= OrderedDict(sorted_list)
    return sorted_dict

    
def export(analysisType,ws, projectId):
    dbdata = Analysis.objects.filter(analysis_exp__project=projectId, analysis_type__field_name=analysisType)
    if (dbdata):
        row_num = 0
        for a in dbdata:
            field_names = [x.name for x in a._meta.local_fields]
            break
        columns = []
        for names in field_names:
            if(names == "analysis_fields"):
                jsonObj = JsonObjField.objects.get(field_name=analysisType)
                jsonFields = orderByNumber(jsonObj.field_set)
                for keys in jsonFields:
                    columns.append((keys, 4000))
            elif(names == "analysis_import"):
                pass
            elif(names == "analysis_hiGlass"):
                pass
            else:
                columns.append((names, 4000))
        
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
    
        for col_num in xrange(len(columns)):
            ws.write(row_num, col_num, columns[col_num][0], font_style)
            # set column width
            ws.col(col_num).width = columns[col_num][1]
    
        font_style = xlwt.XFStyle()
        font_style.alignment.wrap = 1
        
        
        for obj in dbdata:
            row_num += 1
            row = []
            for field in field_names:
                att = getattr(obj, field)
                if(field == "analysis_fields"):
                    json_data = json.loads(att)
                    for keys in jsonFields:
                        json_val = json_data[keys]
                        row.append(json_val)
                elif(field == "analysis_import"):
                    pass
                elif(field == "analysis_hiGlass"):
                    pass
                else:
                    if ((type(att) != int) and (type(att) != str)):
                        att = str(att)
                    row.append(att)
            for col_num in xrange(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)
    else:
        pass
    
         
@login_required 
def exportAnalysis(request):
    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=mymodel.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    
    projectId = request.session['projectId']
    ana = Analysis.objects.filter(analysis_exp__project=projectId)
    
    analysis = list(set([x.analysis_type for x in ana]))
    
    for analysisType in analysis:
        print(analysisType, analysis)
        if str(analysisType) == "Hi-C Analysis":
            ws = wb.add_sheet('Hi-C')
            export(analysisType,ws, projectId)
            
        elif str(analysisType) == "3C Analysis":
            ws = wb.add_sheet('3C')
            export(analysisType,ws, projectId)
             
        elif str(analysisType) == "5C Analysis":
            ws = wb.add_sheet('5C')
            export(analysisType,ws, projectId)
        
        elif str(analysisType) == "CaptureC Analysis":
            ws = wb.add_sheet('CaptureC')
            export(analysisType,ws, projectId)
             
        else:
            pass
        
        
    wb.save(response)
    return response


@login_required 
def exportGEO(request):
    projectId = request.session['projectId']
    prj = Project.objects.get(pk=projectId)
    runUnits = SequencingRun.objects.filter(project=projectId)
    files = SeqencingFile.objects.filter(sequencingFile_exp__project=projectId)
    experiments = Experiment.objects.filter(project=projectId)
    bioSample = Biosample.objects.filter(expBio__project=projectId)
    
    
    title = prj.project_name
    summary = prj.project_notes
    contributor1 = str(prj.project_owner)
    contributor2 = prj.project_contributor.all()
    membersList = []
    for values in contributor2:
        membersList.append(values)
    
 
     
    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=GEO.xlsx'
    file_path_new = WORKSPACEPATH+'/organization/static/siteWide/geo_template_new.xlsx'
 
    wb = load_workbook(file_path_new)
    ws = wb.worksheets[0]
    ws.insert_rows = insert_rows
     
    ws.cell(row=9, column=2).value = title
    ws.cell(row=10, column=2).value = summary
    ws.cell(row=12, column=2).value = contributor1
     
    memberRowNo = 13
    for members in membersList:
        insert_rows(ws, row_idx= memberRowNo, cnt = 1, above=True, copy_style=True)
        ws.cell(row=memberRowNo, column=1).value = "contributor"
        ws.cell(row=memberRowNo, column=2).value = str(members)
         
        memberRowNo +=1
     
    rowNo = ws.max_row
    for i in range (rowNo):
        if((ws.cell(row=i+1, column=1).value)=="Sample name"):
            sampleRowNo = i+2
            break
     
    count = 1
     
         
    for sample in bioSample:
        insert_rows(ws, row_idx= sampleRowNo, cnt = 1, above=False, copy_style=False)
        ws.cell(row=sampleRowNo, column=1).value = "Sample " + str(count)
        ws.cell(row=sampleRowNo, column=2).value = str(sample.biosample_name)
        ws.cell(row=sampleRowNo, column=3).value = str(sample.biosample_biosource.biosource_tissue)
        ws.cell(row=sampleRowNo, column=4).value = str(sample.biosample_individual.individual_type)
        if(sample.protocol):
            #ws.cell(row=sampleRowNo, column=5).value = str(sample.protocol.type)
            ws.cell(row=sampleRowNo, column=6).value = str(sample.protocol.enzyme) 
        ws.cell(row=sampleRowNo, column=7).value = str(sample.biosample_biosource.biosource_cell_line)  
        ws.cell(row=sampleRowNo, column=8).value = str("DNA")
        ws.cell(row=sampleRowNo, column=9).value = str(sample.biosample_biosource.biosource_description)
         
        sampleRowNo += 1
        count += 1
     
    rowNo = ws.max_row
    for i in range (rowNo):
        if((ws.cell(row=i+1, column=1).value)=="RAW FILES"):
            rawFilesRowNo = i+3
            break
     
     
    for file in files:
        insert_rows(ws, row_idx= rawFilesRowNo, cnt = 1, above=True, copy_style=False)
        ws.cell(row=rawFilesRowNo, column=1).value = str(file.sequencingFile_name)
        ws.cell(row=rawFilesRowNo, column=3).value = str(file.sequencingFile_sha256sum)
#         ws.cell(row=rawFilesRowNo, column=5).value = str(file.number_of_reads)
         
        rawFilesRowNo += 1
 
    wb.save(response)
    return response


def initialize(tab,sheetTab):
    file_path_new = WORKSPACEPATH+'/organization/static/siteWide/Metadata_entry_form_V3.xlsx'
    wb = load_workbook(file_path_new)
    sheet = wb.get_sheet_by_name(tab)
    maxCol=get_column_letter(sheet.max_column)
    headers = []
    
    for rowOfCellObjects in sheet['B1':maxCol]:         
        for cellObj in rowOfCellObjects:
            if "*" in cellObj.value:
                headers.append(cellObj.value[1:])
            else:
                headers.append(cellObj.value)
    
    sheetTab.append(headers)
    return (sheetTab)

labName = "dekker-lab:"
def appendPublication(pKey, dcicExcelSheet):
    pub = Publication.objects.get(pk=pKey)
    singlePub = []
    singlePub.append(labName +"Publication_"+str(pub.dcic_alias)+"_"+str(pub.pk))
    singlePub.append(str(pub.publication_title))
    singlePub.append(str(pub.publication_id))
    if(pub.attachment):
        singlePub.append(str(FILEUPLOADPATH)+str(pub.attachment))
    else:
        singlePub.append("")
    if(pub.publication_categories):
        singlePub.append(str(pub.publication_categories))
    else:
        singlePub.append("")
    if(pub.exp_sets_prod_in_pub):
        singlePub.append(str(pub.exp_sets_prod_in_pub))
    else:
        singlePub.append("")
    if(pub.exp_sets_used_in_pub):
        singlePub.append(str(pub.exp_sets_used_in_pub))
    else:
        singlePub.append("")
    if(pub.publication_published_by):
        singlePub.append(str(pub.publication_published_by))
    else:
        singlePub.append("")                   
    dcicExcelSheet['Publication'].append(singlePub)

def appendDocument(pKey, dcicExcelSheet):
    doc = Document.objects.get(pk=pKey)
    singleDocument = []
    singleDocument.append(labName +"Document_"+str(doc.dcic_alias)+"_"+str(doc.pk))
    singleDocument.append(doc.description)
    singleDocument.append(str(doc.type))
    if(doc.attachment):
        singleDocument.append(str(FILEUPLOADPATH)+str(doc.attachment))
    else:
        singleDocument.append("")
    singleDocument.append(str(doc.url))
    if(doc.references):
        singleDocument.append(labName +"Publication_"+str(doc.references.dcic_alias)+"_"+str(doc.references.pk))
        appendPublication(doc.references.pk,dcicExcelSheet)
    else:
        singleDocument.append("")
    dcicExcelSheet['Document'].append(singleDocument)

def appendVendor(pKey,dcicExcelSheet):
    ven = Vendor.objects.get(pk=pKey)
    singleVendor = []
    singleVendor.append(str(ven.dcic_alias))
    singleVendor.append("")
    singleVendor.append(str(ven.vendor_title))
    if(ven.vendor_description != None):
        singleVendor.append(str(ven.vendor_description))
    else:
        singleVendor.append("")
    singleVendor.append(str(ven.vendor_url))
    dcicExcelSheet['Vendor'].append(singleVendor)

def appendEnzyme(pKey,dcicExcelSheet):
    enz = Enzyme.objects.get(pk=pKey)
    singleEnzyme = []
    singleEnzyme.append(labName +"Enzyme_"+str(enz.dcic_alias)+"_"+str(enz.pk))
    singleEnzyme.append(enz.enzyme_name)
    singleEnzyme.append(enz.enzyme_description)
    if(enz.enzyme_vendor):
        singleEnzyme.append(str(enz.enzyme_vendor.dcic_alias))
        appendVendor(enz.enzyme_vendor.pk, dcicExcelSheet)
    else:
        singleEnzyme.append("")
    singleEnzyme.append(enz.enzyme_catalog_number)
    singleEnzyme.append(enz.enzyme_recogSeq)
    singleEnzyme.append(enz.enzyme_siteLen)
    singleEnzyme.append(enz.enzyme_cutPos)
    singleEnzyme.append("")
    if(enz.document):
        singleEnzyme.append(labName +"Document_"+str(enz.document.dcic_alias)+"_"+str(enz.document.pk))
        appendDocument(enz.document.pk, dcicExcelSheet)
    else:
        singleEnzyme.append("")
    singleEnzyme.append(enz.url)
    dcicExcelSheet['Enzyme'].append(singleEnzyme)

def appendImageObjects(pKey,dcicExcelSheet):
    img=ImageObjects.objects.get(pk=pKey)
    singleItem = []
    singleItem.append(labName +"Image_"+str(img.dcic_alias)+"_"+str(img.pk))
    singleItem.append(str(FILEUPLOADPATH)+str(img.imageObjects_images))
    singleItem.append(img.description)
    dcicExcelSheet['Image'].append(singleItem)
    
def appendConstruct(pKey,dcicExcelSheet):
    construct = Construct.objects.get(pk=pKey)
    singleItem = []
    singleItem.append(labName +"Construct_"+str(construct.dcic_alias)+"_"+str(construct.pk))
    singleItem.append(construct.construct_name)
    if(construct.construct_description != None):
        singleItem.append(construct.construct_description)
    else:
        singleItem.append("")
    singleItem.append(str(construct.construct_type))
    if(construct.construct_vendor):
        singleItem.append(str(construct.construct_vendor.dcic_alias))
        appendVendor(construct.construct_vendor.pk,dcicExcelSheet)
    else:
        singleItem.append("")
    singleItem.append(construct.construct_designed_to_Target)
    singleItem.append(construct.construct_insert_sequence)
    if(construct.document):
        singleItem.append(labName +"Document_"+str(construct.document.dcic_alias)+"_"+str(construct.document.pk))
        appendDocument(construct.document.pk, dcicExcelSheet)
    else:
        singleItem.append("")
    singleItem.append(construct.construct_tag)
    singleItem.append(construct.construct_vector_backbone)
    singleItem.append("")
    if(construct.references):
        singleItem.append(labName +"Publication_"+str(construct.references.dcic_alias)+"_"+str(construct.references.pk))
        appendPublication(construct.references.pk,dcicExcelSheet)
    else:
        singleItem.append("")
    singleItem.append(construct.url)
    dcicExcelSheet['Construct'].append(singleItem)

def appendTarget(pKey,dcicExcelSheet):
    target=Target.objects.get(pk=pKey)
    singleItem = []
    singleItem.append(labName +"Target"+str(target.dcic_alias)+"_"+str(target.pk))
    if(target.target_description != None):
        singleItem.append(str(target.target_description))
    else:
        singleItem.append("")
    singleItem.append(str(target.targeted_genes))
    singleItem.append(str(target.targeted_region))
    if(target.references):
        singleItem.append(labName +"Publication_"+str(target.references.dcic_alias)+"_"+str(target.references.pk))
        appendPublication(target.references.pk,dcicExcelSheet)
    else:
        singleItem.append("")
    singleItem.append(str(target.dbxrefs))
    dcicExcelSheet['Target'].append(singleItem)

def appendGenomicRegion(pKey,dcicExcelSheet):          
    genomicRegion = GenomicRegions.objects.get(pk=pKey)
    singleItem = []
    singleItem.append(labName +"GenomicRegion_"+str(genomicRegion.dcic_alias)+"_"+str(genomicRegion.pk))
    if(genomicRegion.genomicRegions_genome_assembly):
        singleItem.append(str(genomicRegion.genomicRegions_genome_assembly))
    else:
        singleItem.append("")
    if(genomicRegion.genomicRegions_chromosome != None):
        singleItem.append(str(genomicRegion.genomicRegions_chromosome))
    else:
        singleItem.append("")
    singleItem.append(genomicRegion.genomicRegions_start_coordinate)
    singleItem.append(genomicRegion.genomicRegions_end_coordinate)
    singleItem.append(genomicRegion.genomicRegions_location_description)
    singleItem.append(genomicRegion.genomicRegions_start_location)
    singleItem.append(genomicRegion.genomicRegions_end_location)
    dcicExcelSheet['GenomicRegion'].append(singleItem)

def appendModification(pKey,dcicExcelSheet):
    modificationObj = Modification.objects.get(pk=pKey)
    singleMod = []
    singleMod.append(labName + "Modification_"+str(modificationObj.dcic_alias)+"_"+str(modificationObj.pk))
    singleMod.append(modificationObj.modification_description)
    if(str(modificationObj.modification_type) != None):
        singleMod.append(str(modificationObj.modification_type))
    else:
        singleMod.append("")
    if(modificationObj.constructs):
        singleMod.append(labName + "Constructs_"+str(modificationObj.constructs.dcic_alias)+"_"+str(modificationObj.constructs.pk))
        appendConstruct(modificationObj.constructs.pk, dcicExcelSheet)
    else:
        singleMod.append("")
    if(modificationObj.modification_vendor):
        singleMod.append(str(modificationObj.modification_vendor.dcic_alias))
        appendVendor(modificationObj.modification_vendor.pk, dcicExcelSheet)
    else:
        singleMod.append("")
    singleMod.append(modificationObj.modification_gRNA)
    if(modificationObj.modification_genomicRegions):
        singleMod.append(labName +"GenomicRegions_"+str(modificationObj.modification_genomicRegions.dcic_alias)+"_"+str(modificationObj.modification_genomicRegions.pk))
        appendGenomicRegion(modificationObj.modification_genomicRegions.pk, dcicExcelSheet)
    else:
        singleMod.append("")
    if(modificationObj.target):
        singleMod.append(labName +"Target_"+str(modificationObj.target.dcic_alias)+"_"+str(modificationObj.target.pk))
        appendTarget(modificationObj.target.pk, dcicExcelSheet)
    else:
        singleMod.append("")
    if(modificationObj.references):
        singleMod.append(labName +"Publication_"+str(modificationObj.references.dcic_alias)+"_"+str(modificationObj.references.pk))
        appendPublication(modificationObj.references.pk,dcicExcelSheet)
    else:
        singleMod.append("")
    singleMod.append(modificationObj.url)
    dcicExcelSheet['Modification'].append(singleMod)


def appendBioRep(expPk,singleExp):
    exp = Experiment.objects.get(pk=expPk)
    expSameBiosample = Experiment.objects.filter(experiment_biosample=exp.experiment_biosample)
    
    expFields=json.loads(exp.experiment_fields)
    bioReplicates = []
    
    for e in expSameBiosample:
        expSameFields=json.loads(e.experiment_fields)
        if(sorted(expSameFields.items()) == sorted(expFields.items())):
            bioReplicates.append(e.pk)
    bio_rep_no = (sorted(bioReplicates)).index(expPk)+1
    singleExp.append(bio_rep_no)
    
def appendTechRep(expPk,singleExp):
    exp = Experiment.objects.get(pk=expPk)
    expSameBiosample = Experiment.objects.filter(experiment_biosample=exp.experiment_biosample, protocol=exp.protocol, experiment_enzyme=exp.experiment_enzyme, type=exp.type)
    
    expFields=json.loads(exp.experiment_fields)
    techReplicates = []
    for e in expSameBiosample:
        expSameFields=json.loads(e.experiment_fields)
        if(sorted(expSameFields.items()) == sorted(expFields.items())):
            techReplicates.append(e.pk)
    
    tech_rep_no = (sorted(techReplicates)).index(expPk)+1
    
    singleExp.append(tech_rep_no)
    
    
def populateDict(request):
    projectId = request.session['projectId']
    bioSample = Biosample.objects.filter(expBio__project=projectId)
    dcicExcelSheet=defaultdict(list)

    tabNames = ("Document","Protocol","Publication","IndividualMouse","IndividualHuman","Vendor","Enzyme","Biosource","Construct","TreatmentRnai",
                "TreatmentChemical","GenomicRegion","Target","Modification","Image","BiosampleCellCulture","Biosample","FileFastq","FileFasta",
                "ExperimentHiC","ExperimentSetReplicate","ExperimentSet")
    
    for tab in tabNames:
        dcicExcelSheet[tab] = initialize(tab, dcicExcelSheet[tab])
    
    
#     ##Experiment sets
#     
#     if (ExperimentSet.objects.filter(project=projectId)):
#         expSets = ExperimentSet.objects.filter(project=projectId)
#         for eSet in expSets:
#             singleItem = []
#             singleItem.append(labName +"ExperimentSet_" +str(eSet.experimentSet_name)+"_"+str(eSet.pk))
#             singleItem.append(str(eSet.experimentSet_description))
#             singleItem.append(str(eSet.experimentSet_type))
#             if(eSet.document):
#                 singleItem.append(labName +"Document_"+str(eSet.document)+"_"+str(eSet.document.pk))
#                 appendDocument(eSet.document.pk, dcicExcelSheet)
#             else:
#                 singleItem.append("")
#             dcicExcelSheet['ExperimentSet'].append(singleItem)
    
                    
    ##Biosample
    for sample in bioSample:
        singleSample = []
        singleSample.append(labName +"Biosample_" +str(sample.dcic_alias)+"_"+str(sample.pk))
        singleSample.append(str(sample.biosample_description))
        singleSample.append(labName +"Biosource_"+str(sample.biosample_biosource.dcic_alias)+"_"+str(sample.biosample_biosource.pk))
        if(sample.protocol):
            singleSample.append(labName +"Protocol_"+str(sample.protocol.dcic_alias)+"_"+str(sample.protocol.pk))
            proto = Protocol.objects.get(pk=sample.protocol.pk)
            singleProtocol = []
            singleProtocol.append(labName +"Protocol_" +str(proto.dcic_alias)+"_"+str(proto.pk))
            singleProtocol.append(proto.description)
            if(proto.attachment):
                singleProtocol.append(str(FILEUPLOADPATH)+str(proto.attachment))
            else:
                singleProtocol.append("")
            dcicExcelSheet['Protocol'].append(singleProtocol)
        else:
            singleSample.append("")
        singleSample.append("")
        singleSample.append("")    
        if(sample.biosample_type):
            singleSample.append(labName +"BiosampleCellCulture_"+str(sample.dcic_alias)+"_"+str(sample.pk))
        else:
            singleSample.append("")
        if(sample.modifications.all()):
            modList = []
            for mod in sample.modifications.all():
                modList.append(labName + "Modification_"+str(mod.dcic_alias)+"_"+str(mod.pk))
                appendModification(mod.pk,dcicExcelSheet)
            singleSample.append(",".join(modList))
        else:
            singleSample.append("")
        
        rnai= []
        chemical=[]
        treatmentList=[]
        if(sample.biosample_TreatmentRnai.all()):
            for rnaiTreat in sample.biosample_TreatmentRnai.all():
                rnai.append(labName + "TreatmentRnai_"+str(rnaiTreat.dcic_alias)+"_"+str(rnaiTreat.pk))
            treatmentList.append(",".join(rnai))
        if(sample.biosample_TreatmentChemical.all()):
            for chemTreat in sample.biosample_TreatmentChemical.all():
                chemical.append(labName + "TreatmentChemical_"+str(chemTreat.dcic_alias)+"_"+str(chemTreat.pk))
            treatmentList.append(",".join(chemical))
        singleSample.append(",".join(treatmentList))
        
        if(sample.references):
            singleSample.append(labName +"Publication_"+str(sample.references.dcic_alias)+"_"+str(sample.references.pk))
            appendPublication(sample.references.pk,dcicExcelSheet)
        else:
            singleSample.append("")
        if(sample.dbxrefs):
            singleSample.append(sample.dbxrefs)
        else:
            singleSample.append("")
        
        dcicExcelSheet['Biosample'].append(singleSample)
        
        ##Biosamplecellculture
        if(sample.biosample_type):
            bcc=json.loads(sample.biosample_fields)
            singleBcc = []
            singleBcc.append(labName +"BiosampleCellCulture_"+str(sample.dcic_alias)+"_"+str(sample.pk))
            singleBcc.append(sample.biosample_description)
            #jsonObj = JsonObjField.objects.get(field_name="BiosampleCellCulture")
            #jsonFields = orderByNumber(jsonObj.field_set)
            #jsonList = jsonFields.items()
            #orders=list(map(lambda k: (int(k[1]['order'])), jsonList))
            singleBcc.append(bcc["culture_start_date"])
            singleBcc.append(bcc["culture_duration"])
            singleBcc.append(bcc["culture_duration_units"])
            singleBcc.append(bcc["culture_harvest_date"])
            singleBcc.append(bcc["differentiation_state"])
            singleBcc.append(bcc["follows_sop"])
            singleBcc.append(bcc["karyotype"])
            if(ImageObjects.objects.filter(bioImg__pk=sample.pk)):
                image=ImageObjects.objects.filter(bioImg__pk=sample.pk)
                for imgs in image:
                    if(imgs.imageObjects_type.choice_name=="karyotype_image"):
                        singleBcc.append(labName +"Image_"+str(imgs.dcic_alias)+"_"+str(imgs.pk))
                        appendImageObjects(imgs.pk,dcicExcelSheet)
                    else:
                        singleBcc.append("")
                    if(imgs.imageObjects_type.choice_name=="morphology_image"):
                        singleBcc.append(labName +"Image_"+str(imgs.dcic_alias)+"_"+str(imgs.pk))
                        appendImageObjects(imgs.pk,dcicExcelSheet)
                    else:
                        singleBcc.append("")
            else:
                singleBcc.append("")
                singleBcc.append("")
            singleBcc.append(bcc["passage_number"])
            if(sample.protocol):
                if(sample.protocol.attachment):
                    singleBcc.append(str(FILEUPLOADPATH)+str(sample.protocol.attachment))
                else:
                    singleBcc.append("")
            else:
                singleBcc.append("")
            
            singleBcc.append("") ##protocol_SOP_deviations
            singleBcc.append("") ##protocol_additional
            
            singleBcc.append(bcc["synchronization_stage"])
            singleBcc.append(bcc["dbxrefs"])
            
#             for keys in jsonFields:
#                 json_val = bcc[keys]
#                 singleBcc.append(json_val)

            dcicExcelSheet['BiosampleCellCulture'].append(singleBcc)
           
            
        ##treatments
        if(sample.biosample_TreatmentRnai):
            treatmentRnais = TreatmentRnai.objects.filter(biosamTreatmentRnai=sample.pk)
            for treatmentRnai in treatmentRnais:
                singleItem = []
                singleItem.append(labName +"TreatmentRnai_"+str(treatmentRnai.dcic_alias)+"_"+str(treatmentRnai.pk))
                singleItem.append(treatmentRnai.treatmentRnai_description)
                if(treatmentRnai.treatmentRnai_type):
                    singleItem.append(str(treatmentRnai.treatmentRnai_type))
                else:
                    singleItem.append("")
                if(treatmentRnai.constructs):
                    singleItem.append(labName +"Constructs_"+str(treatmentRnai.constructs.dcic_alias)+"_"+str(treatmentRnai.constructs.pk))
                    appendConstruct(treatmentRnai.constructs.pk, dcicExcelSheet)
                else:
                    singleItem.append("")
                if(treatmentRnai.treatmentRnai_vendor):
                    singleItem.append(str(treatmentRnai.treatmentRnai_vendor.dcic_alias))
                    appendVendor(treatmentRnai.treatmentRnai_vendor.pk, dcicExcelSheet)
                else:
                    singleItem.append("")
                if(treatmentRnai.treatmentRnai_target):
                    singleItem.append(labName +"Target_"+str(treatmentRnai.treatmentRnai_target.dcic_alias)+"_"+str(treatmentRnai.treatmentRnai_target.pk))
                    appendTarget(treatmentRnai.treatmentRnai_target.pk, dcicExcelSheet)
                else:
                    singleItem.append("")
                if(treatmentRnai.treatmentRnai_nucleotide_seq):
                    singleItem.append(treatmentRnai.treatmentRnai_nucleotide_seq)
                else:
                    singleItem.append("")
                if(treatmentRnai.document):
                    singleItem.append(labName +"Document_"+str(treatmentRnai.document.dcic_alias)+"_"+str(treatmentRnai.document.pk))
                    appendDocument(treatmentRnai.document.pk, dcicExcelSheet)
                else:
                    singleProtocol.append("")
                    
                if(treatmentRnai.references):
                    singleItem.append(labName +"Publication_"+str(treatmentRnai.references.dcic_alias)+"_"+str(treatmentRnai.references.pk))
                    appendPublication(treatmentRnai.references.pk,dcicExcelSheet)
                else:
                    singleItem.append("")
                singleItem.append(treatmentRnai.url)
                dcicExcelSheet['TreatmentRnai'].append(singleItem)
                
        
        if(sample.biosample_TreatmentChemical):
            treatmentChemicals = TreatmentChemical.objects.filter(biosamTreatmentChemical=sample.pk)
            for treatmentChemical in treatmentChemicals:
                singleItem = []
                singleItem.append(labName +"TreatmentChemical_"+str(treatmentChemical.dcic_alias)+"_"+str(treatmentChemical.pk))
                singleItem.append(treatmentChemical.treatmentChemical_description)
                singleItem.append(treatmentChemical.treatmentChemical_chemical)
                if(treatmentChemical.treatmentChemical_concentration != 0):
                    singleItem.append(treatmentChemical.treatmentChemical_concentration)
                else:
                    singleItem.append("")
                if(treatmentChemical.treatmentChemical_concentration_units != None):
                    singleItem.append(str(treatmentChemical.treatmentChemical_concentration_units))
                else:
                    singleItem.append("")
                if(treatmentChemical.treatmentChemical_duration != 0):
                    singleItem.append(str(treatmentChemical.treatmentChemical_duration))
                else:
                    singleItem.append("")
                if(treatmentChemical.treatmentChemical_concentration_units != None):
                    singleItem.append(str(treatmentChemical.treatmentChemical_duration_units))
                else:
                    singleItem.append("")
                if(treatmentChemical.treatmentChemical_temperature != 0.0):
                    singleItem.append(str(treatmentChemical.treatmentChemical_temperature))
                else:
                    singleItem.append("")
                if(treatmentChemical.document):
                    singleItem.append(labName +"Document_"+str(treatmentChemical.document.dcic_alias)+"_"+str(treatmentChemical.document.pk))
                    appendDocument(treatmentChemical.document.pk, dcicExcelSheet)
                else:
                    singleItem.append("")
                if(treatmentChemical.references):
                    singleItem.append(labName +"Publication_"+str(treatmentChemical.references.dcic_alias)+"_"+str(treatmentChemical.references.pk))
                    appendPublication(treatmentChemical.references.pk,dcicExcelSheet)
                else:
                    singleItem.append("")
                dcicExcelSheet['TreatmentChemical'].append(singleItem)
                
        if(Biosource.objects.get(bioSource__pk=sample.pk)):
            biosource = Biosource.objects.get(bioSource__pk=sample.pk)
            singleBio = []
            singleBio.append(labName +"Biosource_"+str(biosource.dcic_alias)+"_"+str(biosource.pk))
            singleBio.append(biosource.biosource_description)
            singleBio.append(str(biosource.biosource_type))
            singleBio.append(str(biosource.biosource_cell_line))
            if(biosource.biosource_cell_line_tier != None):
                singleBio.append(str(biosource.biosource_cell_line_tier))
            else:
                singleBio.append("") 
            ###Standard operation protocol 
#             if(biosource.protocol):
#                 singleBio.append(labName +"Protocol_"+str(biosource.protocol)+"_"+str(biosource.protocol.pk))
#                 proto = Protocol.objects.get(pk=biosource.protocol.pk)
#                 singleProtocol = []
#                 singleProtocol.append(labName +"Protocol_" +str(proto.name)+"_"+str(proto.pk))
#                 singleProtocol.append(proto.description)
#                 if(proto.enzyme):
#                     appendEnzyme(proto.enzyme.pk, dcicExcelSheet)
#                 if(proto.document):
#                     singleProtocol.append(labName +"Document_"+str(proto.document)+"_"+str(proto.document.pk))
#                     appendDocument(proto.document.pk, dcicExcelSheet)
#                 dcicExcelSheet['Protocol'].append(singleProtocol)
#             else:
            singleBio.append("")
                
            if(biosource.biosource_vendor):
                singleBio.append(str(biosource.biosource_vendor.dcic_alias))
                appendVendor(biosource.biosource_vendor.pk, dcicExcelSheet)
            else:
                singleBio.append("")
            
            singleBio.append(biosource.cell_line_termid)
            
            if(biosource.biosource_individual):
                singleBio.append(labName +"Individual_"+str(biosource.biosource_individual.dcic_alias)+"_"+str(biosource.biosource_individual.pk))
                indi = Individual.objects.get(pk=biosource.biosource_individual.pk)
                indiJson = json.loads(indi.individual_fields)
                singleIndi = []
                singleIndi.append(labName +"Individual_"+str(indi.dcic_alias)+"_"+str(indi.pk))
                singleIndi.append(indiJson["age"])
                singleIndi.append(indiJson["age_units"])
                
                if(str(biosource.biosource_individual.individual_type)=="IndividualMouse"):
                    singleIndi.append(indiJson["mouse_life_stage"])
                    singleIndi.append(indiJson["mouse_strain"])
                    singleIndi.append(str(indi.individual_vendor.dcic_alias))
                elif(str(biosource.biosource_individual.individual_type)=="IndividualHuman"):
                    singleIndi.append(indiJson["ethnicity"])
                    singleIndi.append(indiJson["health_status"])
                    singleIndi.append(indiJson["life_stage"])
                
                else:
                    singleIndi.append("")
                    singleIndi.append("")
                    singleIndi.append("")
               
                singleIndi.append(indiJson["sex"])
                
                if(indi.document):
                    singleIndi.append(labName +"Document_"+str(indi.document.dcic_alias)+"_"+str(indi.document.pk))
                    appendDocument(indi.document.pk, dcicExcelSheet)
                else:
                    singleIndi.append("")
                singleIndi.append(indi.url)
                singleIndi.append(indi.dbxrefs)
                
                 
                if(str(biosource.biosource_individual.individual_type)=="IndividualMouse"):
                    dcicExcelSheet['IndividualMouse'].append(singleIndi)
                    
                if(str(biosource.biosource_individual.individual_type)=="IndividualHuman"):
                    dcicExcelSheet['IndividualHuman'].append(singleIndi)
                   
            else:    
                singleBio.append("")
                
            if(biosource.modifications):
                modList = []
                for mod in biosource.modifications.all():
                    modList.append(labName + "Modification_"+str(mod.dcic_alias)+"_"+str(mod.pk))
                    appendModification(mod.pk, dcicExcelSheet)
                singleBio.append(",".join(modList))
            else:
                singleBio.append("")
                
            singleBio.append(biosource.biosource_tissue)    
            
            if(biosource.references):
                singleBio.append(labName +"Publication_"+str(biosource.references.dcic_alias)+"_"+str(biosource.references.pk))
                appendPublication(biosource.references.pk,dcicExcelSheet)
               
            else:
                singleBio.append("")
            singleBio.append(biosource.url)
            dcicExcelSheet['Biosource'].append(singleBio)
    
    experiments = Experiment.objects.filter(project=projectId)
    
    ##Experiments
    for exp in experiments:
        expSet = ExperimentSet.objects.filter(experimentSet_exp=exp)
        if str(exp.type) == "Hi-C Exp Protocol":
            singleExp = []
            singleExp.append(labName +"Experiment_" +str(exp.dcic_alias)+"_"+str(exp.pk))
            singleExp.append(str(exp.experiment_description))
            experiment_set_join= "" 
            replicate_set_join= ""
            if(expSet):
                experiment_set = []
                replicate_set = []
                for eSet in expSet:
                    ExpSet = []
                    ExpSet.append(labName +"ExperimentSet_" +str(eSet.dcic_alias)+"_"+str(eSet.pk))
                    ExpSet.append(str(eSet.description))
                    if(eSet.document):
                        ExpSet.append(labName +"Document_"+str(eSet.document.dcic_alias)+"_"+str(eSet.document.pk))
                        appendDocument(eSet.document.pk, dcicExcelSheet)
                    else:
                        ExpSet.append("")
                    if("replicates" in str(eSet.experimentSet_type)):
                        dcicExcelSheet['ExperimentSetReplicate'].append(ExpSet)
                        experiment_set = []
                        replicate_set.append(labName +"ExperimentSet_"+str(eSet.dcic_alias)+"_"+str(eSet.pk))
                    else:
                        ExpSet.insert(2,str(eSet.experimentSet_type))
                        dcicExcelSheet['ExperimentSet'].append(ExpSet)
                        experiment_set.append(labName +"ExperimentSet_"+str(eSet.dcic_alias)+"_"+str(eSet.pk))
                experiment_set_join= ",".join(experiment_set)    
                replicate_set_join= ",".join(replicate_set)
            
            singleExp.append(replicate_set_join)
            appendBioRep(exp.pk,singleExp)
            appendTechRep(exp.pk,singleExp)
#             singleExp.append("") ####*bio_rep_no
#             singleExp.append("") ####*tec_rep_no 
            singleExp.append(experiment_set_join)
            singleExp.append(labName +"Biosample_" +str(exp.experiment_biosample.dcic_alias)+"_"+str(exp.experiment_biosample.pk))
            
            expFields=json.loads(exp.experiment_fields)
            
            singleExp.append(expFields["experiment_type"])
            singleExp.append(str(exp.biosample_quantity))
            singleExp.append(str(exp.biosample_quantity_units))
            
            singleExp.append(str(expFields["biotin_removed"]))
            singleExp.append(str(expFields["crosslinking_method"]))
            singleExp.append(str(expFields["crosslinking_temperature"]))
            singleExp.append(str(expFields["crosslinking_time"]))
            
            if(exp.experiment_enzyme):
                singleExp.append(labName +"Enzyme_" +str(exp.experiment_enzyme.dcic_alias)+"_"+str(exp.experiment_enzyme.pk))
                appendEnzyme(exp.experiment_enzyme.pk, dcicExcelSheet)
            else:
                singleExp.append("")
                
            singleExp.append(str(expFields["digestion_temperature"]))
            singleExp.append(str(expFields["digestion_time"]))
            singleExp.append(str(expFields["enzyme_lot_number"]))
            singleExp.append(str(expFields["follows_sop"]))
            singleExp.append(str(expFields["average_fragment_size"]))
            singleExp.append(str(expFields["fragment_size_range"]))
            singleExp.append(str(expFields["fragment_size_selection_method"]))
            singleExp.append(str(expFields["fragmentation_method"]))
            singleExp.append(str(expFields["library_preparation_date"]))
            singleExp.append(str(expFields["ligation_temperature"]))
            singleExp.append(str(expFields["ligation_time"]))
            singleExp.append(str(expFields["ligation_volume"]))
            
            if(exp.protocol):
                singleExp.append(labName +"Protocol_" +str(exp.protocol.dcic_alias)+"_"+str(exp.protocol.pk))
                proto = Protocol.objects.get(pk=exp.protocol.pk)
                singleProtocol = []
                singleProtocol.append(labName +"Protocol_" +str(proto.dcic_alias)+"_"+str(proto.pk))
                singleProtocol.append(proto.description)
                if(proto.attachment):
                    singleProtocol.append(str(FILEUPLOADPATH)+str(proto.attachment))
                else:
                    singleProtocol.append("")
                dcicExcelSheet['Protocol'].append(singleProtocol)
            
            singleExp.append("") ###protocol_variation
            singleExp.append(expFields["tagging_method"])
            
            if(SeqencingFile.objects.filter(sequencingFile_exp=exp.pk)):
                files = SeqencingFile.objects.filter(sequencingFile_exp=exp.pk)
                fileList = []
                for f in files:
                    fileList.append(labName +"File"+str(f.dcic_alias)+"_"+str(f.pk))
                    singleFile = []
                    if(str(f.file_format)=="fasta"):
                        singleFile.append(labName +"File_"+str(f.dcic_alias)+"_"+str(f.pk))
                        singleFile.append(str(f.file_format))
                        if(f.file_classification != None):
                            singleFile.append(str(f.file_classification))
                        else:
                            singleFile.append("")
                        if(f.file_format_specifications):
                            singleFile.append(str(f.file_format_specifications))
                        else:
                            singleFile.append("")
                        singleFile.append("")
                        singleFile.append("")
                        singleFile.append(f.dbxrefs)
                        singleFile.append(f.sequencingFile_mainPath)
                        
                        dcicExcelSheet['FileFasta'].append(singleFile)
                    elif(str(f.file_format)=="fastq"):
                        singleFile.append(labName +"File_"+str(f.dcic_alias)+"_"+str(f.pk))
                        singleFile.append(str(f.file_format))
                        singleFile.append(str(f.file_classification))
                        if(f.file_format_specifications):
                            singleFile.append(str(f.file_format_specifications))
                        else:
                            singleFile.append("")
                        if(f.file_barcode):
                            singleFile.append(str(f.file_barcode.barcode_index))
                        else:
                            singleFile.append("")
                        if(f.barcode_in_read):
                            singleFile.append(str(f.barcode_in_read))
                        else:
                            singleFile.append("")
                        if(f.file_barcode):
                            singleFile.append(str(f.file_barcode.barcode_position))
                        else:
                            singleFile.append("")
                        singleFile.append(str(f.flowcell_details_chunk))
                        if(f.sequencingFile_run):
                            singleFile.append(str(f.sequencingFile_run))
                        else:
                            singleFile.append("")
                        singleFile.append(str(f.flowcell_details_lane))
                        if(f.sequencingFile_run.run_sequencing_machine != None):
                            singleFile.append(str(f.sequencingFile_run.run_sequencing_machine))
                        else:
                            singleFile.append("")
                        if(f.sequencingFile_run.run_sequencing_instrument != None):
                            singleFile.append(str(f.sequencingFile_run.run_sequencing_instrument))
                        else:
                            singleFile.append("")
                        singleFile.append(str(f.paired_end))
                        singleFile.append(str(f.read_length))
                        singleFile.append("")
                        singleFile.append("")
                        singleFile.append(f.dbxrefs)
                        singleFile.append(f.sequencingFile_mainPath)
                        
                        dcicExcelSheet['FileFastq'].append(singleFile)
                singleExp.append(",".join(fileList))
            
            else:
                singleExp.append("")
            singleExp.append("")##experiment_relation.relationship_type
            singleExp.append("")##experiment_relation.experiment
            
            if(exp.document):
                singleExp.append(labName +"Document_"+str(exp.document.dcic_alias)+"_"+str(exp.document.pk))
                appendDocument(exp.document.pk, dcicExcelSheet)
            else:
                singleExp.append("")
            
            
            if(exp.references):
                singleExp.append(labName +"Publication_"+str(exp.references.dcic_alias)+"_"+str(exp.references.pk))
                appendPublication(exp.references.pk, dcicExcelSheet)
            else:
                singleExp.append("")
            
            singleExp.append(exp.dbxrefs)
            dcicExcelSheet['ExperimentHiC'].append(singleExp)
        
            
#             jsonObj = JsonObjField.objects.get(field_name="Hi-C Protocol")
#             jsonFields = orderByNumber(jsonObj.field_set)
#             jsonList = jsonFields.items()
#             orders=list(map(lambda k: (int(k[1]['order'])), jsonList))
# #             
#             print(orders)
            
            
#             
#             for keys in jsonFields:
#                 json_val = expFields[keys]
#                 singleItem.append(json_val)
#             
#             print(singleItem)

#             
#             singleItem.insert(11, "")
#             singleItem.insert(20, "")
#             singleItem.insert(21, "")
#             singleItem.insert(23, "")
#             singleItem.insert(24, "")
#             singleItem.insert(25, "")
#             singleItem.insert(26, "")
#             singleItem.insert(27, "")
  
            
    
#     imageObj = ImageObjects.objects.filter(project = projectId)
#     
#     for img in imageObj:
#         singleItem = []
#         singleItem.append(labName +"Images_" +str(img.imageObjects_name)+"_"+str(img.pk))
#         singleItem.append(str(img.imageObjects_images))
#         dcicExcelSheet['Image'].append(singleItem)
#   
#   print(dcicExcelSheet)
    return(dcicExcelSheet)

def duplicates(lst, item):
    return [i for i, x in enumerate(lst) if x == item]

def removeDup(dcicExcelSheet):
    newdcicExcelSheet = defaultdict(list)
    for key, valueList in dcicExcelSheet.items():
        newValueList = []
        pk = []
        for v in valueList:
            if(v[0].split("_")[-1] not in pk):
                newValueList .append(v)
                pk.append(v[0].split("_")[-1])
            else:
                continue
#         pos = dict((x, duplicates(pk, x)) for x in set(pk) if pk.count(x) > 1)    
        newdcicExcelSheet[key] = newValueList
    return(newdcicExcelSheet)
                
                
@login_required 
def exportDCIC(request):
    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=DCIC.xlsx'
    file_path_new = WORKSPACEPATH+'/organization/static/siteWide/Metadata_entry_form_V3.xlsx'
    wb = load_workbook(file_path_new)
    
    dcicExcelSheet = populateDict(request)
    
    dcicExcelSheetOrdered = collections.OrderedDict(dcicExcelSheet)

    dcicExcelSheetDedup = removeDup(dcicExcelSheetOrdered)
    
    for key, valueList in dcicExcelSheetDedup.items():
        ws = wb.get_sheet_by_name(key)
        for r in reversed(list(ws.rows)):
            values = [cell.value for cell in r]
            if any(values):
                maxRow=r[0].row+1
                break
        del valueList[0]
        for v in valueList:
            for i in range(0,len(v)):
                ws.cell(row=maxRow, column=i+2).value = v[i]
            maxRow +=1
    
 
    wb.save(response)
    return response
    
    

    
    
    